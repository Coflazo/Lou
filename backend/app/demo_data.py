from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook

from .config import settings
from .models import Playbook, PlaybookPosition, Proposal, Role


ROOT = Path(__file__).resolve().parents[2]
DEMO_DATA_DIR = ROOT / "demo-data"
XLSX_PATH = Path(settings.DEMO_PLAYBOOK_XLSX)


def ensure_demo_xlsx() -> Path:
    if not XLSX_PATH.exists():
        raise RuntimeError(
            f"Missing demo playbook XLSX at {XLSX_PATH}. "
            "Restore demo-data/siemens-mutual-nda-playbook.xlsx before importing playbooks."
        )
    return XLSX_PATH


def load_demo_playbook() -> Playbook:
    ensure_demo_xlsx()
    wb = load_workbook(XLSX_PATH)
    ws = wb["NDA Playbook"]
    columns = [str(cell.value).strip() for cell in ws[1] if cell.value is not None]
    positions: list[PlaybookPosition] = []
    playbook_id = "pb-mutual-nda"
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        values = {column: "" if value is None else str(value) for column, value in zip(columns, row)}
        topic = column_value(values, "Topic")
        preferred = column_value(values, "Preferred Position")
        fallback = first_column_value(values, ["Fallback Position", "Fallback 1", "Fallback"])
        risk = first_column_value(values, ["Risk", "Red Line", "Deal Breaker"]) or "Medium"
        keywords = first_column_value(values, ["Keywords"]) or infer_keywords(values)
        positions.append(
            PlaybookPosition(
                id=f"pos-{index:02d}",
                playbook_id=playbook_id,
                topic=topic,
                preferred_position=preferred,
                fallback_position=fallback,
                risk=risk,
                keywords=[item.strip().lower() for item in str(keywords).split(";") if item.strip()],
                columns=values,
            )
        )

    return Playbook(
        id=playbook_id,
        slug="mutual-nda",
        name="Siemens Mutual NDA Playbook",
        category="Confidentiality",
        description="Negotiation positions for supplier and partner mutual NDAs.",
        owner="Siemens Legal Ops",
        version=1,
        columns=columns,
        positions=positions,
    )


def column_value(values: dict[str, str], name: str) -> str:
    normalized = normalize_column(name)
    for key, value in values.items():
        if normalize_column(key) == normalized:
            return value
    return ""


def normalize_column(name: str) -> str:
    return "".join(character.lower() for character in name if character.isalnum())


def first_column_value(values: dict[str, str], names: list[str]) -> str:
    for name in names:
        value = column_value(values, name)
        if value:
            return value
    return ""


def infer_keywords(values: dict[str, str]) -> str:
    text = " ".join(values.values())
    words = re.findall(r"[A-Za-z][A-Za-z-]{4,}", text)
    seen = []
    stop = {"position", "fallback", "preferred", "party", "legal", "terms", "clause", "agreement"}
    for word in words:
        lowered = word.lower()
        if lowered not in stop and lowered not in seen:
            seen.append(lowered)
        if len(seen) == 6:
            break
    return ";".join(seen)


def initial_proposals() -> list[Proposal]:
    return [
        Proposal(
            id="prop-demo-residuals",
            playbook_id="pb-mutual-nda",
            topic="Residual Knowledge",
            source="seed",
            proposed_text="Add a senior-review note: residual knowledge clauses require escalation unless they exclude source code, pricing, product plans, and personal data.",
            rationale="Seed proposal for the hackathon approval flow.",
            created_by_role=Role.JUNIOR,
        )
    ]
