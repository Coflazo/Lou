#!/usr/bin/env python3
import sys

if sys.path and sys.path[0] and not sys.path[0].startswith("/"):
    del sys.path[0]

import argparse
import getpass
import json
import os
import re
import time
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import Workbook


SCRIPT_PATH = Path(__file__)
ROOT = SCRIPT_PATH.parents[1] if SCRIPT_PATH.is_absolute() else Path(os.environ.get("PWD", "."))
DEMO_DATA = ROOT / "demo-data"
API_KEYS = ROOT / "api-keys.txt"
DATASET_JSONL = DEMO_DATA / "lou-pioneer-playbook-matrix-50x50.jsonl"
REQUEST_JSON = DEMO_DATA / "pioneer-playbook-matrix-request.json"
RESPONSE_JSON = DEMO_DATA / "pioneer-playbook-matrix-response.json"
WORKBOOK_XLSX = DEMO_DATA / "lou-pioneer-playbook-matrix-50x50.xlsx"
LEGACY_WORKBOOK_XLSX = DEMO_DATA / "siemens-mutual-nda-playbook.xlsx"
PIONEER_CHAT_URL = "https://api.pioneer.ai/v1/chat/completions"
DEFAULT_MODEL = "Qwen/Qwen3-8B"
DEFAULT_SPEC_BATCH_SIZE = 10
ROW_COLUMNS = [
    "Topic",
    "Preferred Position",
    "Fallback 1",
    "Fallback 2",
    "Fallback 3",
    "Red Line",
    "Deal Breaker",
]
KEY_READ_ERRORS: list[str] = []
ROW_RECOVERY_AREAS = [
    "pre-contract diligence and onboarding",
    "commercial pricing and payment operations",
    "data governance and privacy operations",
    "security controls and incident response",
    "AI governance and model-risk controls",
    "IP ownership, licensing, and restrictions",
    "audit, reporting, and records access",
    "termination, transition, and exit support",
    "subcontracting and third-party dependencies",
    "regulatory, export, and public-sector compliance",
    "insurance, liability, indemnity, and remedies",
    "cross-border, localization, and jurisdiction-specific issues",
]


def candidate_roots() -> list[Path]:
    roots: list[Path] = []
    for value in [os.environ.get("LOU_ROOT"), os.environ.get("PWD")]:
        if value:
            roots.append(Path(value))
    try:
        roots.append(Path.cwd())
    except PermissionError:
        pass
    roots.extend([ROOT, SCRIPT_PATH.parents[1], Path(".")])

    seen = set()
    unique = []
    for root in roots:
        key = str(root)
        if key not in seen:
            seen.add(key)
            unique.append(root)
    return unique


def parse_key_line(line: str) -> tuple[str, str] | None:
    stripped = line.strip()
    if not stripped or stripped.startswith("#") or "=" not in stripped:
        return None
    if stripped.startswith("export "):
        stripped = stripped[len("export ") :].strip()
    key, value = stripped.split("=", 1)
    return key.strip(), value.strip().strip("\"'")


def load_keys() -> dict[str, str]:
    keys: dict[str, str] = dict(os.environ)
    for root in candidate_roots():
        path = root / "api-keys.txt"
        try:
            if not path.exists():
                continue
            lines = path.read_text().splitlines()
        except PermissionError as error:
            KEY_READ_ERRORS.append(f"{path}: {error}")
            continue
        for line in lines:
            parsed = parse_key_line(line)
            if parsed:
                key, value = parsed
                keys.setdefault(key, value)
    return keys


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug[:54] or "playbook"


def unique_topic(topic: str, signatures: set[str], playbook: dict[str, str], row_number: int) -> str:
    base = topic.strip() or f"{playbook['name']} Position {row_number}"
    signature = base.casefold()
    if signature not in signatures:
        signatures.add(signature)
        return base

    category = str(playbook.get("category") or "Commercial").strip()
    for suffix in [
        f"{category} Variant {row_number}",
        f"Negotiation Edge Case {row_number}",
        f"Fallback Scenario {row_number}",
        f"Escalation Variant {row_number}",
    ]:
        candidate = f"{base} - {suffix}"
        signature = candidate.casefold()
        if signature not in signatures:
            signatures.add(signature)
            return candidate

    counter = 2
    while True:
        candidate = f"{base} - Variant {row_number}.{counter}"
        signature = candidate.casefold()
        if signature not in signatures:
            signatures.add(signature)
            return candidate
        counter += 1


def parse_json(content: str) -> dict[str, Any]:
    cleaned = content.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, flags=re.S)
        if not match:
            raise
        return json.loads(match.group(0))


def pioneer_chat(messages: list[dict[str, str]], model: str, api_key: str, max_tokens: int = 4096) -> dict[str, Any]:
    payload = {
        "model": model,
        "messages": messages,
        "temperature": 0.82,
        "max_tokens": max_tokens,
    }
    request = urllib.request.Request(
        PIONEER_CHAT_URL,
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={"X-API-Key": api_key, "Content-Type": "application/json"},
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        body = json.loads(response.read().decode("utf-8"))
    content = body["choices"][0]["message"]["content"]
    return {"raw": body, "parsed": parse_json(str(content))}


def generate_playbook_specs(
    count: int,
    spec_batch_size: int,
    model: str,
    api_key: str,
) -> tuple[list[dict[str, str]], dict[str, Any]]:
    validated = []
    names = set()
    attempts = []
    request_index = 0
    empty_attempts = 0
    max_empty_attempts = 8
    max_requests = max(10, count)

    while len(validated) < count:
        request_index += 1
        if request_index > max_requests:
            raise ValueError(f"Pioneer returned only {len(validated)} unique playbooks after {max_requests} requests, expected {count}.")

        current = min(spec_batch_size, count - len(validated))
        messages = [
            {
                "role": "system",
                "content": "Return strict JSON only. No markdown. No extra keys.",
            },
            {
                "role": "user",
                "content": (
                    f"Generate {current} distinct commercial legal playbook specifications. "
                    f"This is request {request_index}. Already accepted {len(validated)} of {count} playbooks. "
                    f"Do not duplicate these accepted playbook names: {json.dumps(sorted(names))}. "
                    "Each must be a different playbook, not just a topic. Cover NDAs, SaaS, DPA, AI vendors, security, "
                    "procurement, services, reseller, IP licensing, employment-adjacent commercial restrictions, cloud, "
                    "support SLAs, privacy, open source, escrow, audit, compliance, payments, termination, liability, "
                    "indemnity, export control, insurance, subcontracting, and public sector. "
                    'Return only {"playbooks":[{"name":"...","category":"...","description":"..."}]}.'
                ),
            },
        ]
        try:
            result = pioneer_chat(messages, model, api_key, max_tokens=8192)
        except json.JSONDecodeError as error:
            attempts.append(
                {
                    "request": request_index,
                    "requested": current,
                    "received": "invalid_json",
                    "accepted": 0,
                    "error": str(error),
                }
            )
            empty_attempts += 1
            if empty_attempts >= max_empty_attempts:
                raise ValueError(
                    f"Pioneer returned malformed playbook specs {empty_attempts} times; "
                    f"accepted {len(validated)} of {count}. Try a smaller --spec-batch-size."
                ) from error
            print(
                f"[pioneer] playbook specs returned malformed JSON; retrying {empty_attempts}/{max_empty_attempts}",
                file=sys.stderr,
            )
            time.sleep(0.25)
            continue
        playbooks = result["parsed"].get("playbooks", [])
        received = len(playbooks) if isinstance(playbooks, list) else "invalid"
        accepted = 0
        if isinstance(playbooks, list):
            for item in playbooks:
                if not isinstance(item, dict):
                    continue
                row = {key: str(item.get(key, "")).strip() for key in ["name", "category", "description"]}
                if not all(row.values()):
                    continue
                signature = row["name"].casefold()
                if signature in names:
                    continue
                names.add(signature)
                validated.append(row)
                accepted += 1
                if len(validated) >= count:
                    break
        attempts.append(
            {
                "request": request_index,
                "requested": current,
                "received": received,
                "accepted": accepted,
                "response": result["parsed"],
            }
        )
        if accepted == 0:
            empty_attempts += 1
            if empty_attempts >= max_empty_attempts:
                raise ValueError(f"Pioneer stopped producing unique playbooks after {len(validated)} of {count}.")
            print(
                f"[pioneer] playbook specs returned no usable items; retrying {empty_attempts}/{max_empty_attempts}",
                file=sys.stderr,
            )
        else:
            empty_attempts = 0
            if accepted < current:
                print(
                    f"[pioneer] accepted {accepted}/{current} playbook specs; requesting the remaining specs",
                    file=sys.stderr,
                )
        time.sleep(0.25)

    return validated[:count], {"attempts": attempts}


def generate_rows_for_playbook(
    playbook: dict[str, str],
    row_count: int,
    batch_size: int,
    model: str,
    api_key: str,
) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    rows: list[dict[str, str]] = []
    batches = []
    signatures = set()
    request_index = 0
    empty_attempts = 0
    max_empty_attempts = 12

    while len(rows) < row_count:
        request_index += 1
        remaining = row_count - len(rows)
        recovery_mode = empty_attempts >= 3
        current = max(batch_size, min(20, remaining * (4 if recovery_mode else 3)))
        existing_topics = [row["Topic"] for row in rows]
        recovery_area = ROW_RECOVERY_AREAS[(request_index - 1) % len(ROW_RECOVERY_AREAS)]
        messages = [
            {
                "role": "system",
                "content": "Return strict JSON only. No markdown. No extra keys.",
            },
            {
                "role": "user",
                "content": (
                    f"Generate {current} rows for this legal playbook: {json.dumps(playbook)}. "
                    f"This is request {request_index}. Already accepted {len(rows)} of {row_count} rows. "
                    f"I still need {remaining} new accepted rows. You may return extra candidates. "
                    f"Do not duplicate any of these accepted topics: {json.dumps(existing_topics)}. "
                    "Each row must contain exactly these string fields: "
                    f"{', '.join(ROW_COLUMNS)}. "
                    "Each row pattern is: one preferred position, three progressively less preferred but acceptable fallbacks, "
                    "one red line requiring escalation, and one deal breaker that is unacceptable. "
                    "Rows must be concrete, varied, concise, and usable by a commercial legal team. "
                    "If this is late in the sequence, invent narrow non-overlapping subtopics, edge cases, exceptions, "
                    "jurisdictional variants, operational obligations, remedies, audit mechanics, notice mechanics, "
                    "or sector-specific positions instead of restating earlier topics. "
                    f"Focus especially on {recovery_area}. "
                    "Use highly specific Topic labels of 5 to 12 words. "
                    'Return only {"items":[{...}]}.'
                ),
            },
        ]
        try:
            result = pioneer_chat(messages, model, api_key, max_tokens=8192)
        except json.JSONDecodeError as error:
            batches.append(
                {
                    "playbook": playbook,
                    "request": request_index,
                    "requested": current,
                    "received": "invalid_json",
                    "accepted": 0,
                    "error": str(error),
                }
            )
            empty_attempts += 1
            if empty_attempts >= max_empty_attempts:
                raise ValueError(
                    f"{playbook['name']}: Pioneer returned malformed row JSON after {len(rows)} of {row_count} rows."
                ) from error
            print(
                f"[pioneer] {playbook['name']} returned malformed row JSON; retrying {empty_attempts}/{max_empty_attempts}",
                file=sys.stderr,
            )
            time.sleep(0.25)
            continue
        items = result["parsed"].get("items", [])
        if not isinstance(items, list):
            rows_value = result["parsed"].get("rows", [])
            items = rows_value if isinstance(rows_value, list) else []
        received = len(items)
        accepted = 0
        for item in items:
            if not isinstance(item, dict):
                continue
            row = {column: str(item.get(column, "")).strip() for column in ROW_COLUMNS}
            if not all(row.values()):
                continue
            row["Topic"] = unique_topic(row["Topic"], signatures, playbook, len(rows) + 1)
            rows.append(row)
            accepted += 1
            if len(rows) >= row_count:
                break
        batches.append(
            {
                "playbook": playbook,
                "request": request_index,
                "requested": current,
                "received": received,
                "accepted": accepted,
                "response": result["parsed"],
            }
        )
        if accepted == 0:
            empty_attempts += 1
            if empty_attempts >= max_empty_attempts:
                raise ValueError(
                    f"{playbook['name']}: Pioneer stopped producing usable rows after {len(rows)} of {row_count} rows."
                )
            print(
                f"[pioneer] {playbook['name']} returned no usable rows; retrying {empty_attempts}/{max_empty_attempts}",
                file=sys.stderr,
            )
        else:
            empty_attempts = 0
            if accepted < current:
                print(
                    f"[pioneer] {playbook['name']} accepted {accepted}/{current}; requesting the remaining rows",
                    file=sys.stderr,
                )
        time.sleep(0.25)
    return rows[:row_count], batches


def write_outputs(playbook_rows: list[tuple[dict[str, str], list[dict[str, str]]]], raw: dict[str, Any]) -> None:
    DEMO_DATA.mkdir(exist_ok=True)
    with DATASET_JSONL.open("w") as file:
        for playbook_index, (playbook, rows) in enumerate(playbook_rows, start=1):
            playbook_id = f"pb-{playbook_index:02d}-{slugify(playbook['name'])}"
            for row_index, row in enumerate(rows, start=1):
                file.write(
                    json.dumps(
                        {
                            "id": f"{playbook_id}-row-{row_index:02d}",
                            "playbook_id": playbook_id,
                            "playbook_name": playbook["name"],
                            "category": playbook["category"],
                            "description": playbook["description"],
                            **row,
                        }
                    )
                    + "\n"
                )

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for _, (playbook, rows) in enumerate(playbook_rows, start=1):
        sheet = wb.create_sheet(title=playbook["name"][:31])
        sheet.append(["Playbook Name", "Category", "Description", *ROW_COLUMNS])
        for row in rows:
            sheet.append([playbook["name"], playbook["category"], playbook["description"], *[row[column] for column in ROW_COLUMNS]])
    wb.save(WORKBOOK_XLSX)

    legacy = Workbook()
    sheet = legacy.active
    sheet.title = "Generated Playbooks"
    sheet.append(["Playbook Name", "Category", "Description", *ROW_COLUMNS])
    for playbook, rows in playbook_rows:
        for row in rows:
            sheet.append([playbook["name"], playbook["category"], playbook["description"], *[row[column] for column in ROW_COLUMNS]])
    legacy.save(LEGACY_WORKBOOK_XLSX)

    REQUEST_JSON.write_text(
        json.dumps(
            {
                "provider": "pioneer",
                "endpoint": PIONEER_CHAT_URL,
                "schema": ["playbook_id", "playbook_name", "category", "description", *ROW_COLUMNS],
                "playbooks": len(playbook_rows),
                "rows_per_playbook": len(playbook_rows[0][1]) if playbook_rows else 0,
            },
            indent=2,
        )
        + "\n"
    )
    RESPONSE_JSON.write_text(json.dumps(raw, indent=2) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate 50 distinct Pioneer playbooks with 50 rows each.")
    parser.add_argument("--playbooks", type=int, default=50)
    parser.add_argument("--rows-per-playbook", type=int, default=50)
    parser.add_argument("--batch-size", type=int, default=5)
    parser.add_argument("--spec-batch-size", type=int, default=DEFAULT_SPEC_BATCH_SIZE)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    args = parser.parse_args()

    if args.spec_batch_size < 1:
        raise SystemExit("--spec-batch-size must be at least 1.")
    if args.batch_size < 1:
        raise SystemExit("--batch-size must be at least 1.")

    api_key = load_keys().get("PIONEER_API_KEY", "")
    if not api_key and sys.stdin.isatty():
        api_key = getpass.getpass("PIONEER_API_KEY: ").strip()
    if not api_key:
        detail = "\n".join(f"  - {error}" for error in KEY_READ_ERRORS)
        suffix = f"\nCould not read key files:\n{detail}" if detail else ""
        raise SystemExit(f"Missing PIONEER_API_KEY. Refusing to generate non-Pioneer fallback data.{suffix}")

    specs, specs_raw = generate_playbook_specs(args.playbooks, args.spec_batch_size, args.model, api_key)
    outputs = []
    raw: dict[str, Any] = {"playbook_specs": specs_raw, "row_batches": []}
    for index, playbook in enumerate(specs, start=1):
        print(f"[pioneer] {index}/{len(specs)} {playbook['name']}", file=sys.stderr)
        rows, batches = generate_rows_for_playbook(playbook, args.rows_per_playbook, args.batch_size, args.model, api_key)
        outputs.append((playbook, rows))
        raw["row_batches"].extend(batches)
        write_outputs(outputs, raw)

    write_outputs(outputs, raw)
    print(
        json.dumps(
            {
                "provider": "pioneer",
                "playbooks": len(outputs),
                "rows": sum(len(rows) for _, rows in outputs),
                "dataset": str(DATASET_JSONL),
                "workbook": str(WORKBOOK_XLSX),
                "app_workbook": str(LEGACY_WORKBOOK_XLSX),
                "request": str(REQUEST_JSON),
                "response": str(RESPONSE_JSON),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
